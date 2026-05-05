"""
Análisis de márgenes v2 — costo desde share.csv columna R.

Columnas de salida:
  sku · cuenta · item_id · title · logistic_type · size_category
  unidades_vendidas · total_revenue_mxn · dias_con_ventas
  primera_venta · ultima_venta
  precio_min · precio_avg · precio_max · precios_distintos · comision_pct
  costo_share
  mb_costo_share_min_pct · mb_costo_share_avg_pct · mb_costo_share_max_pct
  mn_costo_share_min_pct · mn_costo_share_avg_pct · mn_costo_share_max_pct

Run:
    cd /Users/je/dev/kubera
    uv run python -m fulfillment.tests.margin_analysis_v2
"""

import csv
import sys
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

load_dotenv(Path(__file__).resolve().parents[1] / ".env")

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
from fulfillment.restock.db import get_client  # noqa: E402

IVA = 0.16
SHARE_CSV = Path(__file__).resolve().parent / "share.csv"


# ──────────────────────────────────────────────
# 1. Load costo from share.csv (col R = costo)
# ──────────────────────────────────────────────

def load_share_costs() -> dict[str, float]:
    costs: dict[str, float] = {}
    with SHARE_CSV.open(encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            sku = (row.get("SKU") or "").strip()
            raw = (row.get("costo") or "").strip()
            if sku and raw:
                try:
                    costs[sku] = float(raw)
                except ValueError:
                    pass
    return costs


# ──────────────────────────────────────────────
# 2. Fetch daily_sales from Supabase
# ──────────────────────────────────────────────

def fetch_sales() -> list[dict]:
    db = get_client()
    rows = []
    page, size = 0, 1000
    while True:
        batch = (
            db.table("daily_sales")
            .select("date,cuenta,item_id,sku,title,units_sold,revenue,sale_fee")
            .gt("units_sold", 0)
            .range(page * size, (page + 1) * size - 1)
            .execute()
            .data
        )
        rows.extend(batch)
        if len(batch) < size:
            break
        page += 1
    return rows


# ──────────────────────────────────────────────
# 3. Fetch stock metadata from Supabase
# ──────────────────────────────────────────────

def fetch_stock_meta() -> dict[tuple, dict]:
    db = get_client()
    rows = []
    page, size = 0, 1000
    while True:
        batch = (
            db.table("daily_stock")
            .select("sku,cuenta,size_category,logistic_type,date")
            .not_.is_("sku", "null")
            .order("date", desc=True)
            .range(page * size, (page + 1) * size - 1)
            .execute()
            .data
        )
        rows.extend(batch)
        if len(batch) < size:
            break
        page += 1

    meta: dict[tuple, dict] = {}
    for r in rows:
        key = (r["sku"], r["cuenta"])
        if key not in meta:
            meta[key] = {
                "size_category": r.get("size_category") or "—",
                "logistic_type": r.get("logistic_type") or "—",
            }
    return meta


# ──────────────────────────────────────────────
# 4. Aggregate sales + compute margins
# ──────────────────────────────────────────────

def _mb(price, cost) -> float | None:
    if price is None or cost is None or price <= 0:
        return None
    return round((price - cost) / price, 4)


def _mn(price, cost, comision_pct) -> float | None:
    if price is None or cost is None or price <= 0:
        return None
    fees = price * comision_pct * (1 + IVA)
    return round((price - cost - fees) / price, 4)


def build_report(sales, stock_meta, costs) -> list[dict]:
    agg: dict[tuple, dict] = defaultdict(lambda: {
        "title": None,
        "prices": [],
        "comision_pcts": [],
        "total_units": 0,
        "total_revenue": 0.0,
        "total_sale_fee": 0.0,
        "dates": [],
    })

    for r in sales:
        sku = r.get("sku")
        if not sku:
            continue
        key = (sku, r["cuenta"], r["item_id"])
        a = agg[key]
        if not a["title"] and r.get("title"):
            a["title"] = r["title"]

        units = r.get("units_sold") or 0
        rev = float(r.get("revenue") or 0)
        fee = float(r.get("sale_fee") or 0)

        if units > 0 and rev > 0:
            a["prices"].append(round(rev / units, 4))
        if rev > 0 and fee > 0:
            a["comision_pcts"].append(fee / rev)

        a["total_units"] += units
        a["total_revenue"] += rev
        a["total_sale_fee"] += fee
        a["dates"].append(r["date"])

    rows_out = []
    for (sku, cuenta, item_id), a in agg.items():
        prices = sorted(set(round(p, 2) for p in a["prices"]))
        if not prices:
            continue

        precio_min = min(prices)
        precio_avg = round(a["total_revenue"] / a["total_units"], 2) if a["total_units"] else None
        precio_max = max(prices)

        comision_pct = (
            sum(a["comision_pcts"]) / len(a["comision_pcts"])
            if a["comision_pcts"] else
            (a["total_sale_fee"] / a["total_revenue"] if a["total_revenue"] else 0)
        )

        meta = stock_meta.get((sku, cuenta), {})
        costo = costs.get(sku)

        rows_out.append({
            "sku": sku,
            "cuenta": cuenta,
            "item_id": item_id,
            "title": a["title"] or "—",
            "logistic_type": meta.get("logistic_type", "—"),
            "size_category": meta.get("size_category", "—"),
            "unidades_vendidas": a["total_units"],
            "total_revenue_mxn": round(a["total_revenue"], 2),
            "dias_con_ventas": len(set(a["dates"])),
            "primera_venta": min(a["dates"]),
            "ultima_venta": max(a["dates"]),
            "precio_min": precio_min,
            "precio_avg": precio_avg,
            "precio_max": precio_max,
            "precios_distintos": len(prices),
            "comision_pct": round(comision_pct, 4),
            "costo_share": costo,
            "mb_costo_share_min_pct": _mb(precio_min, costo),
            "mb_costo_share_avg_pct": _mb(precio_avg, costo),
            "mb_costo_share_max_pct": _mb(precio_max, costo),
            "mn_costo_share_min_pct": _mn(precio_min, costo, comision_pct),
            "mn_costo_share_avg_pct": _mn(precio_avg, costo, comision_pct),
            "mn_costo_share_max_pct": _mn(precio_max, costo, comision_pct),
        })

    rows_out.sort(key=lambda r: (r["cuenta"], -(r["total_revenue_mxn"] or 0)))
    return rows_out


# ──────────────────────────────────────────────
# 5. Write XLSX
# ──────────────────────────────────────────────

def write_xlsx(report: list[dict], out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Márgenes v2"

    fieldnames = list(report[0].keys())

    PCT_COLS = {
        "comision_pct",
        "mb_costo_share_min_pct", "mb_costo_share_avg_pct", "mb_costo_share_max_pct",
        "mn_costo_share_min_pct", "mn_costo_share_avg_pct", "mn_costo_share_max_pct",
    }
    MXN_COLS = {"total_revenue_mxn", "costo_share", "precio_min", "precio_avg", "precio_max"}

    MARGIN_COLS = [
        "mb_costo_share_min_pct", "mb_costo_share_avg_pct", "mb_costo_share_max_pct",
        "mn_costo_share_min_pct", "mn_costo_share_avg_pct", "mn_costo_share_max_pct",
    ]

    # Header
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, name in enumerate(fieldnames, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, row_data in enumerate(report, 2):
        for col_idx, name in enumerate(fieldnames, 1):
            val = row_data.get(name)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if name in PCT_COLS and val is not None:
                cell.number_format = "0.00%"
            elif name in MXN_COLS and val is not None:
                cell.number_format = '"$"#,##0.00'

    # Color scale: red → white → green on margin columns
    last_row = len(report) + 1
    color_rule = ColorScaleRule(
        start_type="num", start_value=-1,  start_color="F8696B",  # red   ≤ -100%
        mid_type="num",   mid_value=0,     mid_color="FFFFFF",    # white = 0%
        end_type="num",   end_value=0.8,   end_color="63BE7B",    # green ≥ 80%
    )
    for col_name in MARGIN_COLS:
        col_letter = get_column_letter(fieldnames.index(col_name) + 1)
        ws.conditional_formatting.add(f"{col_letter}2:{col_letter}{last_row}", color_rule)

    # Column widths
    widths = {
        "sku": 26, "cuenta": 16, "item_id": 16, "title": 42,
        "logistic_type": 14, "size_category": 12,
        "primera_venta": 13, "ultima_venta": 13,
        "costo_share": 13,
    }
    for col_idx, name in enumerate(fieldnames, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(name, 13)

    ws.freeze_panes = "A2"
    wb.save(out_path)


# ──────────────────────────────────────────────
# 6. Main
# ──────────────────────────────────────────────

def main() -> Path:
    print("Cargando costos desde share.csv...")
    costs = load_share_costs()
    print(f"  {len(costs)} SKUs con costo")

    print("Cargando ventas de Supabase...")
    sales = fetch_sales()
    print(f"  {len(sales)} filas de daily_sales")

    print("Cargando metadata de stock (Supabase)...")
    stock_meta = fetch_stock_meta()

    print("Calculando márgenes...")
    report = build_report(sales, stock_meta, costs)

    now = datetime.now(timezone.utc)
    out_path = Path(__file__).resolve().parent / f"margin_analysis_v2_{now.strftime('%Y%m%d_%H%M')}.xlsx"

    write_xlsx(report, out_path)

    # Summary
    total = len(report)
    con_costo = sum(1 for r in report if r["costo_share"] is not None)
    sin_costo = total - con_costo
    perdida = sum(1 for r in report if r.get("mn_costo_share_avg_pct") is not None and r["mn_costo_share_avg_pct"] < 0)

    print(f"\n{'═'*60}")
    print(f"  Publicaciones analizadas  : {total}")
    print(f"  Con costo en share.csv    : {con_costo}  ({round(con_costo/total*100)}%)")
    print(f"  Sin costo en share.csv    : {sin_costo}")
    print(f"  Con margen neto negativo  : {perdida}")

    print(f"\n  Top 5 mejor margen neto (avg):")
    top = sorted(
        [r for r in report if r.get("mn_costo_share_avg_pct") is not None],
        key=lambda r: r["mn_costo_share_avg_pct"], reverse=True
    )[:5]
    for r in top:
        print(f"    {r['sku']:30s} {r['cuenta']:15s}  mn={r['mn_costo_share_avg_pct']:+.1%}  precio_avg=${r['precio_avg']}")

    print(f"\n  Top 5 peor margen neto (avg):")
    bot = sorted(
        [r for r in report if r.get("mn_costo_share_avg_pct") is not None],
        key=lambda r: r["mn_costo_share_avg_pct"]
    )[:5]
    for r in bot:
        print(f"    {r['sku']:30s} {r['cuenta']:15s}  mn={r['mn_costo_share_avg_pct']:+.1%}  precio_avg=${r['precio_avg']}")

    print(f"\n  XLSX → {out_path}")
    return out_path


if __name__ == "__main__":
    main()
