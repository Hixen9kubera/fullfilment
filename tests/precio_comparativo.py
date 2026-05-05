"""
Comparativo de precios: ML real vs costos_ml DB.

Columnas:
  item_id · sku · tienda · precio_ml · precio_tachado_ml
  precio_sugerido_db · precio_base_db · diferencia_pct

Rojo si diferencia > 1% entre precio_ml y precio_sugerido_db.

Run:
    cd /Users/je/dev/kubera
    uv run python -m fulfillment.tests.precio_comparativo
"""

import os
import sys
from datetime import datetime, timezone
from pathlib import Path

import mysql.connector
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

load_dotenv(Path(__file__).resolve().parents[2] / "sales-dashboard" / ".env")
load_dotenv(Path(__file__).resolve().parents[1] / ".env", override=True)

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
from fulfillment.config import CUENTAS, get_ml_manager  # noqa: E402


# ──────────────────────────────────────────────
# 1. Fetch all active items from ML API
# ──────────────────────────────────────────────

def fetch_ml_prices() -> list[dict]:
    rows = []
    for cuenta in CUENTAS:
        print(f"  Obteniendo items de {cuenta}...")
        ml = get_ml_manager(cuenta)
        user_id = ml.get("/users/me")["id"]
        total = ml.get(
            f"/users/{user_id}/items/search", params={"limit": 1}
        )["paging"]["total"]

        all_ids: list[str] = []
        for offset in range(0, total, 100):
            batch_ids = ml.get(
                f"/users/{user_id}/items/search",
                params={"limit": 100, "offset": offset},
            ).get("results", [])
            all_ids.extend(batch_ids)

        for i in range(0, len(all_ids), 20):
            batch = all_ids[i : i + 20]
            items = ml.get("/items", params={"ids": ",".join(batch)})
            for entry in items:
                b = entry.get("body") or {}
                if b.get("status") not in ("active", "paused"):
                    continue
                attrs = {
                    a["id"]: a.get("value_name")
                    for a in b.get("attributes", [])
                }
                rows.append({
                    "item_id": b.get("id"),
                    "sku": attrs.get("SELLER_SKU") or "—",
                    "tienda": cuenta,
                    "titulo": b.get("title", ""),
                    "status": b.get("status"),
                    "precio_ml": b.get("price"),
                    "precio_tachado_ml": b.get("original_price"),
                })

        print(f"    {cuenta}: {sum(1 for r in rows if r['tienda']==cuenta)} items")
    return rows


# ──────────────────────────────────────────────
# 2. Fetch costos_ml from MySQL
# ──────────────────────────────────────────────

def fetch_db_prices() -> dict[str, dict]:
    conn = mysql.connector.connect(
        host=os.environ.get("DB_HOST", "srv1249.hstgr.io"),
        port=int(os.environ.get("DB_PORT", 3306)),
        user=os.environ["DB_USER"],
        password=os.environ["DB_PASSWORD"],
        database=os.environ.get("DB_NAME", "u531713409_kubera_ml"),
        connect_timeout=10,
    )
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT sku, precio_sugerido, precio_base, descuento_pct, ml_estado, calculado_at
        FROM costos_ml
    """)
    result = {r["sku"]: r for r in cur.fetchall()}
    cur.close()
    conn.close()
    return result


# ──────────────────────────────────────────────
# 3. Build report + XLSX
# ──────────────────────────────────────────────

def main() -> Path:
    print("Cargando precios de MercadoLibre...")
    ml_rows = fetch_ml_prices()

    print("Cargando precios de costos_ml (MySQL)...")
    db_prices = fetch_db_prices()
    print(f"  {len(db_prices)} SKUs en costos_ml")

    # Build comparison rows
    report = []
    for r in ml_rows:
        sku = r["sku"]
        db = db_prices.get(sku, {})

        precio_ml = float(r["precio_ml"]) if r["precio_ml"] is not None else None
        precio_tachado_ml = float(r["precio_tachado_ml"]) if r["precio_tachado_ml"] else None
        precio_sug_db = float(db["precio_sugerido"]) if db.get("precio_sugerido") else None
        precio_base_db = float(db["precio_base"]) if db.get("precio_base") else None

        if precio_ml and precio_sug_db and precio_sug_db > 0:
            diferencia_pct = (precio_ml - precio_sug_db) / precio_sug_db
        else:
            diferencia_pct = None

        report.append({
            "item_id": r["item_id"],
            "sku": sku,
            "tienda": r["tienda"],
            "titulo": r["titulo"],
            "status": r["status"],
            "precio_ml": precio_ml,
            "precio_tachado_ml": precio_tachado_ml,
            "precio_sugerido_db": precio_sug_db,
            "precio_base_db": precio_base_db,
            "diferencia_pct": diferencia_pct,
            "en_costos_ml": "sí" if db else "no",
            "db_estado": db.get("ml_estado", "—"),
            "db_calculado_at": str(db["calculado_at"])[:10] if db.get("calculado_at") else None,
        })

    # Sort: biggest discrepancy first
    report.sort(key=lambda r: abs(r["diferencia_pct"] or 0), reverse=True)

    # ── Write XLSX ────────────────────────────────────────────────
    now = datetime.now(timezone.utc)
    out_path = Path(__file__).resolve().parent / f"precio_comparativo_{now.strftime('%Y%m%d_%H%M')}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Precio Comparativo"

    fieldnames = [
        "item_id", "sku", "tienda", "titulo", "status",
        "precio_ml", "precio_tachado_ml",
        "precio_sugerido_db", "precio_base_db",
        "diferencia_pct",
        "en_costos_ml", "db_estado", "db_calculado_at",
    ]
    headers = [
        "Item ID", "SKU", "Tienda", "Título", "Status",
        "Precio ML", "Tachado ML",
        "Precio Sugerido DB", "Precio Base DB",
        "Diferencia %",
        "En costos_ml", "DB Estado", "DB Calculado",
    ]
    PCT_COLS = {"diferencia_pct"}
    MXN_COLS = {"precio_ml", "precio_tachado_ml", "precio_sugerido_db", "precio_base_db"}

    # Header
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data
    red_fill   = PatternFill("solid", fgColor="F8696B")
    green_fill = PatternFill("solid", fgColor="C6EFCE")

    dif_col_idx = fieldnames.index("diferencia_pct") + 1

    for row_idx, row_data in enumerate(report, 2):
        dif = row_data.get("diferencia_pct")
        row_is_mismatch = dif is not None and abs(dif) > 0.01  # >1% diff

        for col_idx, name in enumerate(fieldnames, 1):
            val = row_data.get(name)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)

            if name in PCT_COLS and val is not None:
                cell.number_format = "0.00%"
                if val < -0.01:
                    cell.fill = red_fill
                elif val > 0.01:
                    cell.fill = red_fill
            elif name in MXN_COLS and val is not None:
                cell.number_format = '"$"#,##0.00'

            # Red entire row for mismatch on key columns
            if row_is_mismatch and name in ("precio_ml", "precio_sugerido_db", "diferencia_pct"):
                cell.fill = red_fill

    # Column widths
    widths = {
        "item_id": 16, "sku": 26, "tienda": 16, "titulo": 42, "status": 10,
        "precio_ml": 13, "precio_tachado_ml": 13,
        "precio_sugerido_db": 18, "precio_base_db": 14,
        "diferencia_pct": 13, "en_costos_ml": 13,
        "db_estado": 11, "db_calculado_at": 14,
    }
    for col_idx, name in enumerate(fieldnames, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(name, 13)

    ws.freeze_panes = "A2"
    wb.save(out_path)

    # ── Summary ───────────────────────────────────────────────────
    total = len(report)
    con_db = sum(1 for r in report if r["en_costos_ml"] == "sí")
    sin_db = total - con_db
    discrepancias = sum(1 for r in report if r.get("diferencia_pct") is not None and abs(r["diferencia_pct"]) > 0.01)

    print(f"\n{'═'*55}")
    print(f"  Publicaciones analizadas   : {total}")
    print(f"  Con precio en costos_ml    : {con_db}")
    print(f"  Sin precio en costos_ml    : {sin_db}")
    print(f"  Con discrepancia > 1%      : {discrepancias}")
    print(f"\n  Top 10 mayor discrepancia:")
    for r in report[:10]:
        d = r.get("diferencia_pct")
        if d is None: continue
        print(f"    {r['sku']:28s} {r['tienda']:15s}  ML=${r['precio_ml']:>8.2f}  DB=${r['precio_sugerido_db'] or 0:>8.2f}  dif={d:+.1%}")
    print(f"\n  XLSX → {out_path}")
    return out_path


if __name__ == "__main__":
    main()
