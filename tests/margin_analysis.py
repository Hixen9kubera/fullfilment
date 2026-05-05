"""
Análisis de márgenes por publicación y tienda.

Cruza:
  - daily_sales (Supabase): precio real de venta, comisión ML, unidades
  - daily_stock (Supabase): size_category, logistic_type
  - scraping_alibaba (MySQL): costo_alibaba (min_usd_mxn) y costo_landed (costo_unitario_cbm × tipo_cambio)

Márgenes calculados:
  - Margen Bruto: sin ningún fee ML
  - Margen Neto:  después de comisión ML + IVA sobre comisión (sin costo envío Full)

Si el SKU tuvo ventas a distintos precios → columnas _min / _avg / _max.

Run:
    cd /Users/je/dev/kubera
    uv run python -m fulfillment.tests.margin_analysis
"""

import csv
import os
import sys
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

import mysql.connector
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill, numbers
from openpyxl.utils import get_column_letter

load_dotenv(Path(__file__).resolve().parents[1] / ".env")

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
from fulfillment.restock.db import get_client  # noqa: E402

IVA = 0.16

# ──────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────

def _pct(value: float, price: float) -> float | None:
    if price <= 0:
        return None
    return round(value / price * 100, 2)


def _margin(price: float, cost: float, comision_pct: float) -> dict:
    """Returns bruto and neto margins (absolute MXN and %) for one price point."""
    mb = price - cost
    comision_mxn = price * comision_pct
    iva_mxn = comision_mxn * IVA
    mn = price - cost - comision_mxn - iva_mxn
    return {
        "mb_mxn": round(mb, 2),
        "mb_pct": _pct(mb, price),
        "mn_mxn": round(mn, 2),
        "mn_pct": _pct(mn, price),
    }


# ──────────────────────────────────────────────
# 1. Fetch daily_sales from Supabase
# ──────────────────────────────────────────────

def fetch_sales() -> list[dict]:
    db = get_client()
    rows = []
    page, size = 0, 1000
    while True:
        batch = (
            db.table("daily_sales")
            .select("date,cuenta,item_id,sku,title,units_sold,revenue,sale_fee")
            .gt("units_sold", 0)
            .range(page * size, (page + 1) * size - 1)
            .execute()
            .data
        )
        rows.extend(batch)
        if len(batch) < size:
            break
        page += 1
    return rows


# ──────────────────────────────────────────────
# 2. Fetch latest size_category + logistic_type from daily_stock
# ──────────────────────────────────────────────

def fetch_stock_meta() -> dict[tuple, dict]:
    """Returns {(sku, cuenta): {size_category, logistic_type, list_price}}"""
    db = get_client()
    rows = []
    page, size = 0, 1000
    while True:
        batch = (
            db.table("daily_stock")
            .select("sku,cuenta,size_category,logistic_type,price,date")
            .not_.is_("sku", "null")
            .order("date", desc=True)
            .range(page * size, (page + 1) * size - 1)
            .execute()
            .data
        )
        rows.extend(batch)
        if len(batch) < size:
            break
        page += 1

    meta: dict[tuple, dict] = {}
    for r in rows:
        key = (r["sku"], r["cuenta"])
        if key not in meta:  # already sorted desc → first = latest
            meta[key] = {
                "size_category": r.get("size_category") or "—",
                "logistic_type": r.get("logistic_type") or "—",
                "list_price": float(r["price"]) if r.get("price") else None,
            }
    return meta


# ──────────────────────────────────────────────
# 3. Fetch costs from MySQL (scraping_alibaba)
# ──────────────────────────────────────────────

def _base_sku(sku: str) -> str:
    """Strip trailing numeric size suffix: ALZ-0128-BLN-NEG-36 → ALZ-0128-BLN-NEG."""
    parts = sku.split("-")
    if parts and parts[-1].isdigit():
        return "-".join(parts[:-1])
    return sku


def fetch_alibaba_costs() -> dict[str, dict]:
    """Returns {sku: {costo_alibaba_mxn, costo_landed_mxn}}.

    Both min_usd_mxn and costo_unitario_cbm are stored in USD; multiply by
    tipo_cambio to convert to MXN.
    """
    conn = mysql.connector.connect(
        host=os.environ.get("DB_HOST", "srv1249.hstgr.io"),
        port=int(os.environ.get("DB_PORT", 3306)),
        user=os.environ["DB_USER"],
        password=os.environ["DB_PASSWORD"],
        database=os.environ.get("DB_NAME", "u531713409_kubera_ml"),
        connect_timeout=10,
    )
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT sku,
               min_usd_mxn        AS alibaba_usd,
               costo_unitario_cbm AS landed_usd,
               tipo_cambio
        FROM scraping_alibaba
        WHERE costo_unitario_cbm IS NOT NULL OR min_usd_mxn IS NOT NULL
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    costs: dict[str, dict] = {}
    for r in rows:
        tc = float(r["tipo_cambio"] or 19)
        # Both columns are already in MXN — no tc conversion needed
        ali_mxn = float(r["alibaba_usd"]) if r["alibaba_usd"] else None
        cbm_mxn = float(r["landed_usd"]) if r["landed_usd"] else None
        costs[r["sku"]] = {
            "costo_alibaba_mxn": ali_mxn,
            "costo_landed_mxn": cbm_mxn,
            "tipo_cambio": tc,
        }
    return costs


# ──────────────────────────────────────────────
# 4. Aggregate and compute margins
# ──────────────────────────────────────────────

def build_report(sales: list[dict], stock_meta: dict, costs: dict[str, dict]) -> list[dict]:
    # Group daily_sales by (sku, cuenta, item_id)
    agg: dict[tuple, dict] = defaultdict(lambda: {
        "title": None,
        "prices": [],          # effective unit price per day
        "comision_pcts": [],   # sale_fee / revenue per day (when revenue > 0)
        "total_units": 0,
        "total_revenue": 0.0,
        "total_sale_fee": 0.0,
        "dates": [],
    })

    for r in sales:
        sku = r.get("sku")
        if not sku:
            continue
        key = (sku, r["cuenta"], r["item_id"])
        a = agg[key]
        if not a["title"] and r.get("title"):
            a["title"] = r["title"]

        units = r.get("units_sold") or 0
        rev = float(r.get("revenue") or 0)
        fee = float(r.get("sale_fee") or 0)

        if units > 0 and rev > 0:
            a["prices"].append(round(rev / units, 4))
        if rev > 0 and fee > 0:
            a["comision_pcts"].append(fee / rev)

        a["total_units"] += units
        a["total_revenue"] += rev
        a["total_sale_fee"] += fee
        a["dates"].append(r["date"])

    rows_out = []
    for (sku, cuenta, item_id), a in agg.items():
        prices = sorted(set(round(p, 2) for p in a["prices"]))
        if not prices:
            continue

        precio_min = min(prices)
        precio_avg = round(a["total_revenue"] / a["total_units"], 2) if a["total_units"] else None
        precio_max = max(prices)
        has_variacion = (precio_min != precio_max)

        comision_pct = (
            round(sum(a["comision_pcts"]) / len(a["comision_pcts"]), 4)
            if a["comision_pcts"] else
            round(a["total_sale_fee"] / a["total_revenue"], 4) if a["total_revenue"] else 0
        )

        meta = stock_meta.get((sku, cuenta), {})
        cost_data = costs.get(sku) or costs.get(_base_sku(sku), {})
        costo_alibaba = cost_data.get("costo_alibaba_mxn")
        costo_landed = cost_data.get("costo_landed_mxn")

        did_perc = (
            round((costo_landed - costo_alibaba) / costo_alibaba, 4)
            if costo_alibaba and costo_landed else None
        )

        def mb_pct(price, cost):
            if price is None or cost is None or price <= 0:
                return None
            return round((price - cost) / price, 4)  # decimal for Excel % format

        row: dict = {
            "sku": sku,
            "cuenta": cuenta,
            "item_id": item_id,
            "title": a["title"] or "—",
            "logistic_type": meta.get("logistic_type", "—"),
            "size_category": meta.get("size_category", "—"),
            "unidades_vendidas": a["total_units"],
            "total_revenue_mxn": round(a["total_revenue"], 2),
            "dias_con_ventas": len(set(a["dates"])),
            "primera_venta": min(a["dates"]),
            "ultima_venta": max(a["dates"]),
            "precio_min": precio_min,
            "precio_avg": precio_avg,
            "precio_max": precio_max,
            "precios_distintos": len(prices) if has_variacion else 1,
            "comision_pct": round(comision_pct * 100, 2),
            "costo_alibaba_mxn": costo_alibaba,
            "costo_landed_mxn": costo_landed,
            "did_perc": did_perc,
            "mb_alibaba_min_pct": mb_pct(precio_min, costo_alibaba),
            "mb_alibaba_avg_pct": mb_pct(precio_avg, costo_alibaba),
            "mb_alibaba_max_pct": mb_pct(precio_max, costo_alibaba),
            "mb_landed_min_pct": mb_pct(precio_min, costo_landed),
            "mb_landed_avg_pct": mb_pct(precio_avg, costo_landed),
            "mb_landed_max_pct": mb_pct(precio_max, costo_landed),
            "datos_sospechosos": (
                "sí" if (costo_landed and precio_avg and costo_landed > precio_avg * 3) else "no"
            ),
        }

        rows_out.append(row)

    rows_out.sort(key=lambda r: (r["cuenta"], -(r["total_revenue_mxn"] or 0)))
    return rows_out


# ──────────────────────────────────────────────
# 5. Main
# ──────────────────────────────────────────────

def main() -> Path:
    print("Cargando ventas de Supabase...")
    sales = fetch_sales()
    print(f"  {len(sales)} filas de daily_sales")

    print("Cargando metadata de stock (Supabase)...")
    stock_meta = fetch_stock_meta()
    print(f"  {len(stock_meta)} combinaciones sku/cuenta")

    print("Cargando costos de scraping_alibaba (MySQL)...")
    costs = fetch_alibaba_costs()
    print(f"  {len(costs)} SKUs con costo")

    print("Calculando márgenes...")
    report = build_report(sales, stock_meta, costs)

    now = datetime.now(timezone.utc)
    out_path = Path(__file__).resolve().parent / f"margin_analysis_{now.strftime('%Y%m%d_%H%M')}.xlsx"

    if not report:
        print("Sin datos para exportar.")
        return out_path

    # ── Build Excel workbook ──────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Márgenes"

    fieldnames = list(report[0].keys())

    PCT_COLS  = {"did_perc", "mb_alibaba_min_pct", "mb_alibaba_avg_pct", "mb_alibaba_max_pct",
                 "mb_landed_min_pct", "mb_landed_avg_pct", "mb_landed_max_pct", "comision_pct"}
    MXN_COLS  = {"total_revenue_mxn", "costo_alibaba_mxn", "costo_landed_mxn",
                 "precio_min", "precio_avg", "precio_max"}

    # Header row
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, name in enumerate(fieldnames, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, row_data in enumerate(report, 2):
        for col_idx, name in enumerate(fieldnames, 1):
            val = row_data.get(name)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if name in PCT_COLS and val is not None:
                cell.number_format = "0.00%"
            elif name in MXN_COLS and val is not None:
                cell.number_format = '"$"#,##0.00'

    # Conditional formatting on the 6 margin columns (red–white–green)
    margin_col_names = [
        "mb_alibaba_min_pct", "mb_alibaba_avg_pct", "mb_alibaba_max_pct",
        "mb_landed_min_pct",  "mb_landed_avg_pct",  "mb_landed_max_pct",
    ]
    last_row = len(report) + 1
    color_rule = ColorScaleRule(
        start_type="num",  start_value=-1,   start_color="F8696B",   # red   (≤ -100%)
        mid_type="num",    mid_value=0,       mid_color="FFFFFF",     # white (0%)
        end_type="num",    end_value=0.8,     end_color="63BE7B",     # green (≥ 80%)
    )
    for col_name in margin_col_names:
        if col_name in fieldnames:
            col_letter = get_column_letter(fieldnames.index(col_name) + 1)
            col_range = f"{col_letter}2:{col_letter}{last_row}"
            ws.conditional_formatting.add(col_range, color_rule)

    # Also color did_perc: white at 0, orange at high values
    if "did_perc" in fieldnames:
        col_letter = get_column_letter(fieldnames.index("did_perc") + 1)
        ws.conditional_formatting.add(
            f"{col_letter}2:{col_letter}{last_row}",
            ColorScaleRule(
                start_type="num", start_value=0,   start_color="FFFFFF",
                end_type="num",   end_value=5,     end_color="F4721E",
            ),
        )

    # Column widths
    col_widths = {
        "sku": 26, "cuenta": 16, "item_id": 16, "title": 40,
        "logistic_type": 14, "size_category": 12,
        "primera_venta": 13, "ultima_venta": 13,
    }
    for col_idx, name in enumerate(fieldnames, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(name, 13)

    # Freeze header
    ws.freeze_panes = "A2"

    wb.save(out_path)

    # ── Summary ───────────────────────────────────────────────────
    total_skus = len(report)
    con_costo  = sum(1 for r in report if r.get("costo_alibaba_mxn") is not None)
    sin_costo  = total_skus - con_costo

    perdida_landed  = [r for r in report if r.get("mb_landed_avg_pct")  is not None and r["mb_landed_avg_pct"]  < 0]
    perdida_alibaba = [r for r in report if r.get("mb_alibaba_avg_pct") is not None and r["mb_alibaba_avg_pct"] < 0]

    print(f"\n{'═'*60}")
    print(f"  Publicaciones analizadas       : {total_skus}")
    print(f"  Con costo Alibaba              : {con_costo}  ({round(con_costo/total_skus*100)}%)")
    print(f"  Sin costo Alibaba              : {sin_costo}")
    print(f"  Margen bruto negativo (landed) : {len(perdida_landed)}")
    print(f"  Margen bruto negativo (alibaba): {len(perdida_alibaba)}")

    print(f"\n  Top 5 mejor margen bruto (landed, precio avg):")
    top = sorted(
        [r for r in report if r.get("mb_landed_avg_pct") is not None],
        key=lambda r: r["mb_landed_avg_pct"], reverse=True
    )[:5]
    for r in top:
        print(f"    {r['sku']:30s} {r['cuenta']:15s}  mb={r['mb_landed_avg_pct']:+.1%}  precio_avg=${r['precio_avg']}")

    print(f"\n  Top 5 peor margen bruto (landed, precio avg):")
    bot = sorted(
        [r for r in report if r.get("mb_landed_avg_pct") is not None],
        key=lambda r: r["mb_landed_avg_pct"]
    )[:5]
    for r in bot:
        print(f"    {r['sku']:30s} {r['cuenta']:15s}  mb={r['mb_landed_avg_pct']:+.1%}  precio_avg=${r['precio_avg']}")

    print(f"\n  XLSX → {out_path}")
    return out_path


if __name__ == "__main__":
    main()
